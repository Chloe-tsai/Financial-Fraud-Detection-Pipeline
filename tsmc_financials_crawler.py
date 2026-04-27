"""
台積電 (2330) 四大財務報表爬蟲 + Excel 輸出
資料來源：公開資訊觀測站 MOPS (mops.twse.com.tw)

安裝依賴：
    pip install requests pandas openpyxl beautifulsoup4 lxml

執行：
    python tsmc_financials_crawler.py

輸出：
    台積電四大財報_2023_2024.xlsx
"""

import time
from io import StringIO

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 設定 ───────────────────────────────────────────────────────────────────

COMPANY_ID       = "2330"
TARGET_YEARS_ROC = [112, 113]          # 民國年 112=2023, 113=2024
OUTPUT_FILE      = "台積電四大財報_2023_2024.xlsx"
SLEEP_SEC        = 2.0                 # 每次請求間隔（秒），避免被擋

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Referer": "https://mops.twse.com.tw/",
    "Accept-Language": "zh-TW,zh;q=0.9",
}

MOPS_ENDPOINTS = {
    "損益表":     "https://mops.twse.com.tw/mops/web/ajax_t164sb04",
    "資產負債表": "https://mops.twse.com.tw/mops/web/ajax_t164sb03",
    "現金流量表": "https://mops.twse.com.tw/mops/web/ajax_t164sb05",
    "股東權益表": "https://mops.twse.com.tw/mops/web/ajax_t164sb06",
}

# ── 爬蟲 ───────────────────────────────────────────────────────────────────

def fetch_mops(url: str, year_roc: int, season: int) -> pd.DataFrame | None:
    """
    POST 到 MOPS，取回指定公司、年度、季別的財報 DataFrame。
    year_roc : 民國年（如 113）
    season   : 1~4
    回傳最大 <table> 的解析結果；失敗回傳 None。
    """
    payload = {
        "encodeURIComponent": "1",
        "step":       "1",
        "firstin":    "1",
        "off":        "1",
        "keyword4":   "",
        "code1":      "",
        "TYPEK2":     "",
        "checkbtn":   "",
        "queryName":  "co_id",
        "inpuType":   "co_id",
        "TYPEK":      "all",
        "isnew":      "false",
        "co_id":      COMPANY_ID,
        "year":       str(year_roc),
        "season":     str(season),
    }
    try:
        resp = requests.post(url, data=payload, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        resp.encoding = "utf-8"
    except requests.RequestException as e:
        print(f"    ✗ 連線失敗：{e}")
        return None

    soup   = BeautifulSoup(resp.text, "lxml")
    tables = soup.find_all("table")
    if not tables:
        print("    ✗ 找不到資料表（該季可能尚未公告）")
        return None

    main = max(tables, key=lambda t: len(t.find_all("tr")))
    try:
        df = pd.read_html(StringIO(str(main)), header=0)[0]
    except Exception as e:
        print(f"    ✗ 解析失敗：{e}")
        return None

    df.dropna(how="all", inplace=True)
    df.dropna(axis=1, how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    cols    = list(df.columns)
    cols[0] = "科目"
    df.columns = cols
    return df


def crawl_all() -> dict:
    """
    爬取所有年份、所有季別的四大報表。
    回傳結構：{ "Q1 2023": {"損益表": df, ...}, ... }
    """
    results: dict = {}
    total = len(TARGET_YEARS_ROC) * 4 * len(MOPS_ENDPOINTS)
    done  = 0

    for year_roc in TARGET_YEARS_ROC:
        ad_year = year_roc + 1911
        for season in range(1, 5):
            label = f"Q{season} {ad_year}"
            print(f"\n▶ {label}")
            results[label] = {}
            for stmt_name, url in MOPS_ENDPOINTS.items():
                print(f"  抓取 {stmt_name}...", end=" ", flush=True)
                df = fetch_mops(url, year_roc, season)
                results[label][stmt_name] = df
                done += 1
                if df is not None:
                    print(f"✓ ({len(df)} 列)")
                time.sleep(SLEEP_SEC)

    success = sum(
        1 for p in results.values() for df in p.values() if df is not None
    )
    print(f"\n共 {total} 次請求，成功 {success} 筆。")
    return results


# ── Excel 樣式 ─────────────────────────────────────────────────────────────

_thin  = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

HDR_FONT  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDR_FILL  = PatternFill("solid", start_color="1F4E79")
BOLD_FONT = Font(name="Arial", size=10, bold=True, color="000000")
BODY_FONT = Font(name="Arial", size=10, color="000000")
GRAY_FONT = Font(name="Arial", size=9,  color="595959")
ALT_FILL  = PatternFill("solid", start_color="EEF4FB")
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center")
RIGHT     = Alignment(horizontal="right",  vertical="center")

NUM_FMT = '#,##0;(#,##0);"-"'
PCT_FMT = '0.0%;(0.0%);"-"'


# ── Excel 建構 ─────────────────────────────────────────────────────────────

def build_sheet(wb: Workbook, sheet_name: str, data: dict):
    ws      = wb.create_sheet(title=sheet_name)
    periods = list(data.keys())
    ncols   = 1 + len(periods)

    # 收集所有科目（保持首次出現順序）
    all_items: list[str] = []
    seen: set[str] = set()
    for period in periods:
        df = data[period].get(sheet_name)
        if df is None:
            continue
        for item in df["科目"].astype(str).tolist():
            item = item.strip()
            if item and item not in seen:
                all_items.append(item)
                seen.add(item)

    if not all_items:
        ws["A1"] = "（無資料）"
        return

    # ── 標題列 ──
    ws.row_dimensions[1].height = 30
    hc = ws.cell(row=1, column=1, value="科目")
    hc.font = HDR_FONT; hc.fill = HDR_FILL; hc.border = BORDER; hc.alignment = LEFT
    ws.column_dimensions["A"].width = 38

    for ci, period in enumerate(periods, start=2):
        c = ws.cell(row=1, column=ci, value=period)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER; c.alignment = CENTER
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # ── 建立快速查找表 period -> {科目: 數值} ──
    lookup: dict[str, dict[str, object]] = {}
    for period in periods:
        df = data[period].get(sheet_name)
        lookup[period] = {}
        if df is None:
            continue
        val_cols = [c for c in df.columns if c != "科目"]
        for row in df.itertuples(index=False):
            item = str(row[0]).strip()
            for v in row[1:]:
                try:
                    num = float(
                        str(v).replace(",", "")
                              .replace("－", "-")
                              .replace("—", "")
                              .replace("−", "-")
                    )
                    lookup[period][item] = num
                    break
                except (ValueError, TypeError):
                    continue

    # ── 資料列 ──
    PCT_KEYWORDS = ("率", "比率", "佔", "占", "每股")
    for row_i, item in enumerate(all_items, start=2):
        alt  = (row_i % 2 == 0)
        fill = ALT_FILL if alt else PatternFill()
        ws.row_dimensions[row_i].height = 15

        lc = ws.cell(row=row_i, column=1, value=item)
        lc.font = BOLD_FONT; lc.fill = fill; lc.border = BORDER; lc.alignment = LEFT

        is_pct  = any(k in item for k in PCT_KEYWORDS)
        is_elem = "每股" in item

        for ci, period in enumerate(periods, start=2):
            val = lookup[period].get(item)
            dc  = ws.cell(row=row_i, column=ci, value=val)
            dc.fill = fill; dc.border = BORDER; dc.alignment = RIGHT
            if val is not None:
                dc.font = BODY_FONT
                if is_pct:
                    dc.value         = val / 100.0
                    dc.number_format = PCT_FMT
                else:
                    dc.number_format = NUM_FMT

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"

    note_row = len(all_items) + 3
    nc = ws.cell(row=note_row, column=1,
                 value="單位：千元新台幣（另有標示者除外）｜資料來源：公開資訊觀測站 mops.twse.com.tw")
    nc.font = GRAY_FONT; nc.alignment = LEFT


def build_cover(wb: Workbook):
    ws = wb.create_sheet(title="封面", index=0)
    ws.column_dimensions["A"].width = 65

    ad_years = [y + 1911 for y in TARGET_YEARS_ROC]
    yr_range = f"{ad_years[0]} Q1 – {ad_years[-1]} Q4"

    rows = [
        ("台積電 (2330) 四大財務報表",
         Font(name="Arial", size=20, bold=True, color="1F4E79")),
        (f"資料期間：{yr_range}（各季）",
         Font(name="Arial", size=12, color="595959")),
        ("單位：千元新台幣（另有標示者除外）",
         Font(name="Arial", size=11, color="595959")),
        ("資料來源：公開資訊觀測站 (mops.twse.com.tw)",
         Font(name="Arial", size=10, color="595959")),
        ("包含報表：損益表 ／ 資產負債表 ／ 現金流量表 ／ 股東權益表",
         Font(name="Arial", size=10, color="595959")),
        ("", Font(name="Arial", size=10)),
        ("注意事項：",
         Font(name="Arial", size=10, bold=True, color="595959")),
        ("• 各季資料為當季單季數；現金流量表為累計數。",
         Font(name="Arial", size=10, color="595959")),
        ("• 比率類科目已自動轉換為百分比格式。",
         Font(name="Arial", size=10, color="595959")),
        ("• 若某季顯示空白，表示 MOPS 當時尚未公告或解析失敗。",
         Font(name="Arial", size=10, color="595959")),
    ]
    for ri, (text, font) in enumerate(rows, start=1):
        c = ws.cell(row=ri, column=1, value=text)
        c.font = font; c.alignment = LEFT
        ws.row_dimensions[ri].height = 24
    ws.row_dimensions[1].height = 44


def export_excel(data: dict):
    wb = Workbook()
    wb.remove(wb.active)
    build_cover(wb)
    for stmt in ["損益表", "資產負債表", "現金流量表", "股東權益表"]:
        print(f"  寫入 sheet：{stmt}")
        build_sheet(wb, stmt, data)
    wb.save(OUTPUT_FILE)
    print(f"\n✅ 已儲存至 {OUTPUT_FILE}")


# ── 主程式 ─────────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  台積電 (2330) 四大財報爬蟲")
    print(f"  目標年度：{[y+1911 for y in TARGET_YEARS_ROC]}")
    print(f"  輸出檔案：{OUTPUT_FILE}")
    print("=" * 55)
    data = crawl_all()
    print("\n正在產生 Excel...")
    export_excel(data)


if __name__ == "__main__":
    main()



